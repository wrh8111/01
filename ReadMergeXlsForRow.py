import openpyxl
import re

# 横向读取表格数据，默认从第二行，第一列开始读取。
def readMergeXlsForRow(startRow = 2,startCol = 1):
    work_book = openpyxl.load_workbook("test.xlsx")
    sheet1 = work_book.worksheets[0]

    cell_list = []
    for row in range(startRow,sheet1.max_row+1):
        for col in range(startCol, sheet1.max_column+1):
            if sheet1.cell(row, col).value != None:
                cell_list.append(sheet1.cell(row, col).value)
    return cell_list

#去除单个数据中的手机号码
def getWithoutPhoneData(data):
    return re.sub("(13[0-9]|14[15679]|15[0-3,5-9]|166|17[0-8]|18[0-9]|19[89])\\d{8}", "", data)

#获取数据中的手机号码，如果没找到手机号码，返回"没找到电话号码"
def getPhoneDate(data):
    result = "没找到电话号码"
    matchObj = re.search("(13[0-9]|14[15679]|15[0-3,5-9]|166|17[0-8]|18[0-9]|19[89])\\d{8}", data)
    if matchObj:
        result = matchObj.group()
    return result


if __name__ == "__main__":
    li = readMergeXlsForRow()
    for i in range(len(li)):
        # li[i] = re.sub("(13[0-9]|14[15679]|15[0-3,5-9]|166|17[0-8]|18[0-9]|19[89])\\d{8}", "", li[i])
        print(li[i])
        print(getWithoutPhoneData(li[i]))
        print(getPhoneDate(li[i]))
